import openpyxl

def parse_excel(filepath: str, filename: str) -> list[dict]:
    """Excelをシート・テーブル単位でチャンク化"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))

        if not rows:
            continue

        # ヘッダー行をセクション名に使用
        section = f"{sheet_name}：{rows[0]}" if rows else sheet_name
        text = "\n".join(rows)

        chunks.append({
            "filename": filename,
            "page": None,
            "section": section,
            "text": text,
        })

    return chunks
